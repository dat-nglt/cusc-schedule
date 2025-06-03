import json
import pandas as pd
import os
import glob
from datetime import datetime, timedelta
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import (
    PatternFill, Border, Side, Alignment, 
    Font, GradientFill, Color
)
from openpyxl.utils import get_column_letter

INPUT_DIR = os.path.join(os.path.dirname(__file__), '../../data/output/weekly')
OUTPUT_EXCEL = 'output/full_timetable.xlsx'
MAPPING_FILE = os.path.join(os.path.dirname(__file__), 'mapping', 'subject_names.json')
LOGO_PATH = os.path.join(os.path.dirname(__file__), 'logo.png')

# Tải dữ liệu mapping
with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
    SUBJECT_NAMES = json.load(f)

# --- Các hàm hỗ trợ ---
def get_time_range(slot_id):
    time_map = {
        'S1': ('07:00', '09:00'),
        'S2': ('09:00', '11:00'),
        'C1': ('13:00', '15:00'),
        'C2': ('15:00', '17:00'),
        'T1': ('17:30', '19:30'),
        'T2': ('19:30', '21:30')
    }
    start, end = time_map.get(slot_id, ('', ''))
    return f"{start}-{end}"

def convert_day(day_abbr):
    days_map = {
        'Mon': 'Thứ 2',
        'Tue': 'Thứ 3',
        'Wed': 'Thứ 4',
        'Thu': 'Thứ 5',
        'Fri': 'Thứ 6',
        'Sat': 'Thứ 7',
        'Sun': 'Chủ nhật'
    }
    return days_map.get(day_abbr, day_abbr)

def create_cell_content(session):
    subject_name = SUBJECT_NAMES.get(session['subject'], session['subject'])
    session_type = ' (LT)' if session['type'] == 'theory' else ' (TH)'
    lecturer = session.get('lecturer', '')
    room = session.get('room', '')
    return f"{subject_name}{session_type}\nGV: {lecturer}\nPhòng: {room}"

# --- Hàm định dạng Excel nâng cao ---
def apply_professional_formatting(ws, week_num, start_date=None):
    # Thiết lập style
    header_fill = PatternFill(
        start_color="1F497D", 
        end_color="1F497D", 
        fill_type="solid"
    )
    header_font = Font(
        name='Arial', 
        size=12, 
        bold=True, 
        color="FFFFFF"
    )
    time_fill = PatternFill(
        start_color="D9E1F2", 
        fill_type="solid"
    )
    day_fill = PatternFill(
        start_color="E2EFDA", 
        fill_type="solid"
    )
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_aligned = Alignment(
        horizontal='center', 
        vertical='center', 
        wrap_text=True
    )
    
    # Áp dụng định dạng cho tất cả ô
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = center_aligned
    
    # Định dạng header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Định dạng cột thời gian
    for cell in ws['B']:
        if cell.row > 1:
            cell.fill = time_fill
    
    # Định dạng các cột ngày
    day_columns = {'D', 'E', 'F', 'G', 'H', 'I', 'J'}
    for col in day_columns:
        for cell in ws[col]:
            if cell.row == 1:
                cell.fill = day_fill
                cell.font = Font(bold=True)
    
    # Merge ô cho các lớp
    current_class = None
    start_row = 0
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1), 2):
        class_name = row[0].value
        if class_name != current_class:
            if current_class is not None:
                ws.merge_cells(f"A{start_row}:A{idx-1}")
            current_class = class_name
            start_row = idx
    
    if current_class is not None:
        ws.merge_cells(f"A{start_row}:A{ws.max_row}")
    
    # Điều chỉnh độ rộng cột
    for col in range(1, ws.max_column + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 5)
        ws.column_dimensions[column].width = adjusted_width
    
    # Thêm tiêu đề
    ws.insert_rows(1)
    ws.merge_cells(f"A1:{get_column_letter(ws.max_column)}1")
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"THỜI KHÓA BIỂU - TUẦN {week_num}"
    title_cell.font = Font(name='Arial', size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Thêm chân trang
    ws.cell(row=ws.max_row+1, column=1).value = "Ghi chú: LT = Lý thuyết, TH = Thực hành"
    ws.cell(row=ws.max_row+1, column=1).value = f"Ngày tạo: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

# --- Hàm xử lý chính ---
def process_week(week_data, week_num, start_date=None):
    rows = []
    
    # Tạo tiêu đề ngày với ngày cụ thể
    date_headers = {}
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        for i, day in enumerate(['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']):
            current_date = start_date + timedelta(days=i)
            date_headers[day] = {
                'name': convert_day(day),
                'date': current_date.strftime('%d/%m')
            }
    
    # Xử lý từng lớp
    for class_id, sessions in week_data['schedule'].items():
        if not sessions:
            continue
            
        # Nhóm session theo khung giờ
        time_slots = sorted(set(session['slot'] for session in sessions))
        
        for slot in time_slots:
            slot_sessions = [s for s in sessions if s['slot'] == slot]
            row = {
                'THỜI GIAN': get_time_range(slot),
                'Class': class_id,
                'Lab': '',
                'Ghi chú': '',
                'SL sinh viên': '',
                'Giáo viên': ''
            }
            
            # Thêm cột ngày
            for day_abbr in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']:
                day_sessions = [s for s in slot_sessions if s['day'] == day_abbr]
                
                if date_headers:
                    day_name = f"{date_headers[day_abbr]['name']}\n{date_headers[day_abbr]['date']}"
                else:
                    day_name = convert_day(day_abbr)
                
                if day_sessions:
                    row[day_name] = '\n\n'.join(create_cell_content(s) for s in day_sessions)
                else:
                    row[day_name] = ''
            
            rows.append(row)
    
    # Tạo DataFrame
    columns = ['THỜI GIAN', 'Class', 'Lab'] 
    if date_headers:
        columns += [f"{date_headers[d]['name']}\n{date_headers[d]['date']}" for d in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']]
    else:
        columns += [convert_day(d) for d in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']]
    
    columns += ['Ghi chú', 'SL sinh viên', 'Giáo viên']
    
    return pd.DataFrame(rows, columns=columns)

def export_to_excel():
    # Tạo thư mục output
    os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)
    
    # Lấy file JSON
    json_files = glob.glob(os.path.join(INPUT_DIR, '*.json'))
    json_files.sort(key=lambda x: int(x.split('_')[-2]) if 'week_' in x else 0)
    
    if not json_files:
        print("⚠️ Không tìm thấy file JSON nào")
        return
    
    # Tạo workbook
    wb = Workbook()
    wb.remove(wb.active)  # Xóa sheet mặc định
    
    # Xác định ngày bắt đầu học kỳ
    start_date = datetime.now().strftime('%Y-%m-%d')
    
    for i, json_file in enumerate(json_files):
        try:
            # Trích xuất số tuần
            filename = os.path.basename(json_file)
            week_num = 1
            if 'week_' in filename:
                week_num = int(filename.split('_')[1])
            
            # Đọc dữ liệu
            with open(json_file, 'r', encoding='utf-8') as f:
                week_data = json.load(f)
            
            print(f"🎨 Đang tạo định dạng đẹp cho tuần {week_num}...")
            
            # Xử lý dữ liệu
            week_df = process_week(week_data, week_num, start_date)
            
            # Cập nhật ngày
            if start_date:
                start_date = (datetime.strptime(start_date, '%Y-%m-%d') + timedelta(weeks=1)).strftime('%Y-%m-%d')
            
            # Tạo sheet mới
            sheet_name = f"Tuần {week_num}"
            ws = wb.create_sheet(title=sheet_name)
            
            # Ghi dữ liệu vào sheet
            for r_idx, row in enumerate(dataframe_to_rows(week_df, index=False, header=True), 1):
                ws.append(row)
            
            # Áp dụng định dạng chuyên nghiệp
            apply_professional_formatting(ws, week_num)
            
        except Exception as e:
            print(f"❌ Lỗi khi xử lý {json_file}: {str(e)}")
    
    # Lưu workbook
    wb.save(OUTPUT_EXCEL)
    print(f"✅ Xuất file Excel chuyên nghiệp thành công: {OUTPUT_EXCEL}")

if __name__ == "__main__":
    export_to_excel()