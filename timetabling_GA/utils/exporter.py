from datetime import datetime, timedelta
import openpyxl, os
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from collections import defaultdict


def export_lecturer_view_to_excel(lecturer_semester_view: dict, processed_data, output_folder: str = "results"):
    """
    Xuất lịch trình học kỳ của từng giảng viên ra một tệp Excel duy nhất.
    Mỗi sheet trong tệp Excel sẽ đại diện cho lịch của một tuần cụ thể.

    Args:
        lecturer_semester_view (dict): Dữ liệu lịch trình học kỳ của giảng viên.
                                       Ví dụ: {lecturer_id: {week_num: [lesson_list, ...]}}
        processed_data (DataProcessor): Đối tượng chứa dữ liệu đã được tiền xử lý (ví dụ: thông tin giảng viên, môn học, slot thời gian).
        output_folder (str): Thư mục để lưu tệp Excel đầu ra. Mặc định là "results".
    """
    if not lecturer_semester_view:
        print("Không có dữ liệu lịch trình giảng viên để xuất. Đã dừng thao tác.")
        return

    # Ánh xạ tên ngày tiếng Anh sang tiếng Việt để hiển thị trên Excel
    days_of_week_map = {"Mon": "Thứ 2", "Tue": "Thứ 3", "Wed": "Thứ 4", "Thu": "Thứ 5", "Fri": "Thứ 6", "Sat": "Thứ 7", "Sun": "CN"}
    
    # Lấy danh sách các ngày trong tuần được sử dụng trong lịch trình (ví dụ: ["Mon", "Tue", ...])
    schedule_days_eng = processed_data.data.get('days_of_week', ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"])
    
    # Tạo tiêu đề cột cho các ngày trong tuần (ví dụ: "Thứ 2", "Thứ 3", ...)
    day_column_headers = [days_of_week_map.get(day_eng, day_eng) for day_eng in schedule_days_eng]
    
    # Tạo ánh xạ từ slot_id sang thông tin chi tiết của slot (start_time, end_time)
    slot_info_map = {s['slot_id']: s for s in processed_data.data['time_slots']}
    
    # Tiêu đề chính của bảng trong Excel
    excel_main_headers = ["Slot Thời Gian", "Giảng Viên"] + day_column_headers

    # --- Định nghĩa Style cho Excel ---
    # Font
    header_font = Font(bold=True, color="FFFFFF", name="Times New Roman", size=11)
    sheet_title_font = Font(bold=True, size=16, name="Times New Roman")
    slot_time_font = Font(bold=True, name="Times New Roman", size=11)
    university_header_font = Font(name="Times New Roman", size=12, bold=True)
    university_subheader_font = Font(name="Times New Roman", size=11)
    university_contact_font = Font(name="Times New Roman", size=10)

    # Fill (màu nền)
    header_fill = PatternFill(start_color="4A86E8", end_color="4A86E8", fill_type="solid") # Màu xanh cho tiêu đề
    theory_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid") # Màu xanh nhạt cho tiết lý thuyết
    practice_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid") # Màu vàng nhạt cho tiết thực hành

    # Alignment (căn chỉnh)
    cell_alignment_wrapped_center = Alignment(wrap_text=True, horizontal='center', vertical='center')
    header_alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Border (đường viền)
    thin_border_side = Side(style='thin', color="000000")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Thông tin header của trường đại học
    university_header_lines = [
        "TRUNG TÂM CÔNG NGHỆ PHẦN MỀM ĐẠI HỌC CẦN THƠ",
        "CANTHO UNIVERSITY SOFTWARE CENTER",
        "Khu III, Đại học Cần Thơ - 01 Lý Tự Trọng, TP. Cần Thơ - Tel: 0292.3731072 & Fax: 0292.3731071 - Email: cusc@ctu.edu.vn"
    ]

    # Xác định số tuần tối đa có dữ liệu lịch trình
    max_weeks = 0
    if lecturer_semester_view:
        # Lấy tất cả các số tuần có dữ liệu từ lecturer_semester_view
        all_weeks_with_data = {week_num for lecturer_data in lecturer_semester_view.values() 
                               for week_num in lecturer_data.keys()}
        if all_weeks_with_data:
            max_weeks = max(all_weeks_with_data)

    if max_weeks == 0:
        print("Không tìm thấy dữ liệu lịch trình cho bất kỳ tuần nào. Đã dừng xuất file.")
        return

    # Tạo một workbook (tệp Excel) mới
    workbook_lecturer = openpyxl.Workbook()
    is_first_sheet = True # Cờ để kiểm tra xem đây có phải sheet đầu tiên không (sheet mặc định)

    # Lặp qua từng tuần để tạo các sheet tương ứng
    for week_num_display in range(1, max_weeks + 1):
        if is_first_sheet:
            sheet = workbook_lecturer.active # Sử dụng sheet mặc định
            sheet.title = f"GV_Tuan_{week_num_display}"
            is_first_sheet = False
        else:
            # Tạo sheet mới cho các tuần tiếp theo
            sheet = workbook_lecturer.create_sheet(title=f"GV_Tuan_{week_num_display}")

        # --- Ghi thông tin Header của Trường Đại học ---
        current_row_for_header = 1
        # Dòng 1: Tên trung tâm bằng tiếng Việt
        sheet.cell(row=current_row_for_header, column=1, value=university_header_lines[0]).font = university_header_font
        sheet.merge_cells(start_row=current_row_for_header, start_column=1, end_row=current_row_for_header, end_column=len(excel_main_headers))
        sheet.cell(row=current_row_for_header, column=1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[current_row_for_header].height = 25
        current_row_for_header += 1
        
        # Dòng 2: Tên trung tâm bằng tiếng Anh
        sheet.cell(row=current_row_for_header, column=1, value=university_header_lines[1]).font = university_subheader_font
        sheet.merge_cells(start_row=current_row_for_header, start_column=1, end_row=current_row_for_header, end_column=len(excel_main_headers))
        sheet.cell(row=current_row_for_header, column=1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[current_row_for_header].height = 20
        current_row_for_header += 1
        
        # Dòng 3: Thông tin liên hệ
        sheet.cell(row=current_row_for_header, column=1, value=university_header_lines[2]).font = university_contact_font
        sheet.merge_cells(start_row=current_row_for_header, start_column=1, end_row=current_row_for_header, end_column=len(excel_main_headers))
        sheet.cell(row=current_row_for_header, column=1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[current_row_for_header].height = 20
        current_row_for_header += 1
        
        # Tạo một hàng trống nhỏ để phân tách
        sheet.row_dimensions[current_row_for_header].height = 10
        current_row_for_header += 1

        # --- Ghi tiêu đề của sheet (LỊCH DẠY GIẢNG VIÊN - TUẦN X) ---
        title_row = current_row_for_header
        sheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=len(excel_main_headers))
        title_cell = sheet.cell(row=title_row, column=1, value=f"LỊCH DẠY GIẢNG VIÊN - TUẦN {week_num_display}")
        title_cell.font = sheet_title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[title_row].height = 30
        
        current_excel_row = title_row + 2 # Bắt đầu hàng dữ liệu sau tiêu đề và một hàng trống
        
        # --- Ghi các header chính của bảng (Slot Thời Gian, Giảng Viên, Thứ 2, ...) ---
        for col_idx, header_text in enumerate(excel_main_headers, start=1):
            cell = sheet.cell(row=current_excel_row, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment_center
            cell.border = cell_border
        sheet.row_dimensions[current_excel_row].height = 35 # Chiều cao hàng header
        current_excel_row += 1 # Chuyển xuống hàng tiếp theo để ghi dữ liệu

        # --- Thu thập và sắp xếp dữ liệu cho tuần hiện tại ---
        all_lessons_this_week = []
        for lecturer_id, weekly_data in lecturer_semester_view.items():
            if week_num_display in weekly_data:
                # Thêm lecturer_id vào mỗi lesson để dễ dàng nhóm và sắp xếp
                for lesson in weekly_data[week_num_display]:
                    lesson_copy = lesson.copy() # Tạo bản sao để tránh sửa đổi dữ liệu gốc
                    lesson_copy['lecturer_id'] = lecturer_id
                    all_lessons_this_week.append(lesson_copy)

        # Nếu không có lịch dạy nào cho tuần này, ghi thông báo và chuyển sang tuần tiếp theo
        if not all_lessons_this_week:
            empty_msg_cell = sheet.cell(row=current_excel_row, column=1, value="Không có giảng viên nào có lịch dạy trong tuần này.")
            sheet.merge_cells(start_row=current_excel_row, start_column=1, end_row=current_excel_row, end_column=len(excel_main_headers))
            empty_msg_cell.alignment = Alignment(horizontal="center")
            empty_msg_cell.font = Font(italic=True, name="Times New Roman")
            
            # Áp dụng border cho các ô trong hàng thông báo rỗng
            for empty_col_idx in range(1, len(excel_main_headers) + 1):
                sheet.cell(row=current_excel_row, column=empty_col_idx).border = cell_border
            continue # Chuyển sang tuần tiếp theo

        # Sắp xếp các tiết học: Ưu tiên theo slot thời gian, sau đó theo ID giảng viên
        # processed_data.slot_order_map phải được đảm bảo có sẵn để sắp xếp đúng thứ tự slot
        all_lessons_this_week.sort(key=lambda x: (processed_data.slot_order_map.get(x['slot_id'], 99), x['lecturer_id']))

        # Tạo một lưới lịch trình tạm thời để dễ dàng truy cập dữ liệu khi ghi vào Excel
        # Cấu trúc: schedule_grid[(slot_id, lecturer_id)][day_of_week_english] = lesson_data
        schedule_grid = defaultdict(lambda: defaultdict(dict))
        for lesson in all_lessons_this_week:
            slot_id = lesson['slot_id']
            lecturer_id = lesson['lecturer_id']
            day = lesson['day'] # 'day' ở đây là tên ngày tiếng Anh (Mon, Tue,...)
            schedule_grid[(slot_id, lecturer_id)][day] = lesson
        
        # Sắp xếp các khóa của schedule_grid để đảm bảo thứ tự hiển thị trong Excel
        sorted_schedule_keys = sorted(schedule_grid.keys(), 
                                      key=lambda x: (processed_data.slot_order_map.get(x[0], 99), x[1]))

        # --- Đổ dữ liệu lịch trình vào sheet ---
        previous_slot_id = None # Dùng để hợp nhất các ô "Slot Thời Gian"
        for slot_id, lecturer_id in sorted_schedule_keys:
            lecturer_name = processed_data.lecturer_map.get(lecturer_id, {}).get('lecturer_name', lecturer_id)
            
            # Ghi thông tin Slot Thời Gian và Giảng Viên
            if slot_id != previous_slot_id:
                # Nếu slot_id thay đổi, ghi slot mới
                slot_start_time = slot_info_map.get(slot_id,{}).get('start', '')
                slot_end_time = slot_info_map.get(slot_id,{}).get('end', '')
                slot_cell = sheet.cell(row=current_excel_row, column=1, value=f"{slot_id}\n({slot_start_time}-{slot_end_time})")
                slot_cell.font = slot_time_font
                slot_cell.alignment = cell_alignment_wrapped_center
                sheet.row_dimensions[current_excel_row].height = 60 # Tăng chiều cao hàng để chứa nội dung
            
            lect_cell = sheet.cell(row=current_excel_row, column=2, value=lecturer_name)
            lect_cell.alignment = cell_alignment_wrapped_center
            
            # Ghi dữ liệu cho từng ngày trong tuần
            for col_idx, day_eng in enumerate(schedule_days_eng, start=3):
                lesson_content = schedule_grid[(slot_id, lecturer_id)].get(day_eng)
                cell = sheet.cell(row=current_excel_row, column=col_idx)
                
                if lesson_content:
                    # Lấy tên môn học và chuyển đổi loại tiết học sang tiếng Việt
                    subject_details = processed_data.subject_map.get(lesson_content['subject_id'], {})
                    subject_name = subject_details.get('name', lesson_content['subject_id'])
                    lesson_type_vi = "Lý thuyết" if lesson_content['lesson_type'] == "theory" else "Thực hành"
                    
                    # Định dạng chuỗi hiển thị nội dung tiết học
                    lesson_content_str = (
                        f"Lớp: {lesson_content['class_id']}\n"
                        f"Môn: {subject_name}\n"
                        f"Phòng: {lesson_content['room_id']}\n"
                        f"({lesson_type_vi})"
                    )
                    cell.value = lesson_content_str
                    # Tô màu nền dựa trên loại tiết học
                    cell.fill = theory_fill if lesson_content['lesson_type'] == "theory" else practice_fill
                else:
                    cell.value = "" # Để trống nếu không có tiết học
                
                cell.alignment = cell_alignment_wrapped_center
                cell.border = cell_border
            
            # Áp dụng border cho cột Slot Thời Gian và Giảng Viên
            sheet.cell(row=current_excel_row, column=1).border = cell_border
            sheet.cell(row=current_excel_row, column=2).border = cell_border
            
            current_excel_row += 1 # Di chuyển xuống hàng tiếp theo
            previous_slot_id = slot_id # Cập nhật slot_id trước đó để hợp nhất

        # --- Hợp nhất các ô "Slot Thời Gian" ---
        # Bắt đầu từ hàng đầu tiên chứa dữ liệu (sau header và title)
        current_row_for_merge = title_row + 3 
        for ts_info in processed_data.data['time_slots']: # Lặp qua từng slot thời gian định nghĩa
            slot_id_to_merge = ts_info['slot_id']
            start_merge_row = current_row_for_merge # Hàng bắt đầu của vùng cần hợp nhất
            
            # Duyệt xuống dưới cho đến khi gặp một slot_id khác hoặc hết dữ liệu
            while current_row_for_merge <= sheet.max_row and \
                  sheet.cell(row=current_row_for_merge, column=1).value and \
                  str(sheet.cell(row=current_row_for_merge, column=1).value).startswith(slot_id_to_merge):
                current_row_for_merge += 1
            
            # Nếu có nhiều hơn một hàng cho cùng một slot_id, thực hiện hợp nhất
            if current_row_for_merge - start_merge_row > 1:
                sheet.merge_cells(start_row=start_merge_row, start_column=1, end_row=current_row_for_merge - 1, end_column=1)

        # --- Điều chỉnh độ rộng cột ---
        sheet.column_dimensions[get_column_letter(1)].width = 18 # Cột "Slot Thời Gian"
        sheet.column_dimensions[get_column_letter(2)].width = 25 # Cột "Giảng Viên"
        for i in range(3, len(excel_main_headers) + 1):
            sheet.column_dimensions[get_column_letter(i)].width = 30 # Các cột ngày trong tuần

    # --- Lưu tệp Excel ---
    # Tạo thư mục đầu ra nếu chưa tồn tại
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        
    output_filename = os.path.join(output_folder, "TKB_HocKy_GiangVien.xlsx")
    
    # Xóa tệp cũ nếu tồn tại để tránh lỗi ghi đè
    if os.path.exists(output_filename):
        try: 
            os.remove(output_filename)
            print(f"Đã xóa tệp cũ: '{output_filename}'.")
        except OSError as e:
            print(f"Lỗi khi xóa tệp Excel giảng viên cũ '{output_filename}': {e}. Vui lòng đóng tệp nếu đang mở.")
            return

    try:
        workbook_lecturer.save(output_filename)
        print(f"🎉 Đã tạo thành công tệp Excel lịch trình giảng viên (tất cả các tuần) tại: '{output_filename}'")
    except OSError as e:
        print(f"Lỗi khi lưu tệp Excel '{output_filename}': {e}. Vui lòng đảm bảo tệp không bị mở bởi ứng dụng khác.")
    except Exception as e:
        print(f"Một lỗi không mong muốn đã xảy ra khi lưu tệp Excel: {e}")
          
def export_room_view_to_excel(room_semester_view: dict, processed_data, output_folder: str = "results"):
    """
    Xuất lịch sử dụng phòng học ra một tệp Excel duy nhất.
    Mỗi sheet trong tệp Excel sẽ đại diện cho lịch của một tuần cụ thể.

    Args:
        room_semester_view (dict): Dữ liệu lịch trình học kỳ của phòng học.
                                   Ví dụ: {room_id: {week_num: [lesson_list, ...]}}
        processed_data (DataProcessor): Đối tượng chứa dữ liệu đã được tiền xử lý (ví dụ: thông tin giảng viên, môn học, slot thời gian).
        output_folder (str): Thư mục để lưu tệp Excel đầu ra. Mặc định là "results".
    """
    if not room_semester_view:
        print("Không có dữ liệu lịch trình phòng học để xuất. Đã dừng thao tác.")
        return

    # Ánh xạ tên ngày tiếng Anh sang tiếng Việt để hiển thị trên Excel
    days_of_week_map = {"Mon": "Thứ 2", "Tue": "Thứ 3", "Wed": "Thứ 4", "Thu": "Thứ 5", "Fri": "Thứ 6", "Sat": "Thứ 7", "Sun": "CN"}
    
    # Lấy danh sách các ngày trong tuần được sử dụng trong lịch trình (ví dụ: ["Mon", "Tue", ...])
    schedule_days_eng = processed_data.data.get('days_of_week', ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"])
    
    # Tạo tiêu đề cột cho các ngày trong tuần (ví dụ: "Thứ 2", "Thứ 3", ...)
    day_column_headers = [days_of_week_map.get(day_eng, day_eng) for day_eng in schedule_days_eng]
    
    # Tạo ánh xạ từ slot_id sang thông tin chi tiết của slot (start_time, end_time)
    slot_info_map = {s['slot_id']: s for s in processed_data.data.get('time_slots', [])}
    
    # Lấy thông tin chi tiết về các slot thời gian
    time_slots_info = processed_data.data.get('time_slots', [])
    
    # Tiêu đề chính của bảng trong Excel
    excel_main_headers = ["Slot Thời Gian", "Phòng"] + day_column_headers
    
    # --- Định nghĩa Style cho Excel ---
    # Font
    header_font = Font(bold=True, color="FFFFFF", name="Times New Roman", size=11)
    sheet_title_font = Font(bold=True, size=16, name="Times New Roman")
    slot_time_font = Font(bold=True, name="Times New Roman", size=11)
    university_header_font = Font(name="Times New Roman", size=12, bold=True)
    university_subheader_font = Font(name="Times New Roman", size=11)
    university_contact_font = Font(name="Times New Roman", size=10)

    # Fill (màu nền)
    header_fill = PatternFill(start_color="6AA84F", end_color="6AA84F", fill_type="solid") # Màu xanh lá cây cho tiêu đề
    theory_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid") # Màu xanh nhạt cho tiết lý thuyết
    practice_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid") # Màu vàng nhạt cho tiết thực hành

    # Alignment (căn chỉnh)
    cell_alignment_wrapped_center = Alignment(wrap_text=True, horizontal='center', vertical='center')
    header_alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Border (đường viền)
    thin_border_side = Side(style='thin', color="000000")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Thông tin header của trường đại học
    university_header_lines = [
        "TRUNG TÂM CÔNG NGHỆ PHẦN MỀM ĐẠI HỌC CẦN THƠ",
        "CANTHO UNIVERSITY SOFTWARE CENTER",
        "Khu III, Đại học Cần Thơ - 01 Lý Tự Trọng, TP. Cần Thơ - Tel: 0292.3731072 & Fax: 0292.3731071 - Email: cusc@ctu.edu.vn"
    ]

    # Xác định số tuần tối đa có dữ liệu lịch trình
    max_weeks = 0
    if room_semester_view:
        # Lấy tất cả các số tuần có dữ liệu từ room_semester_view
        all_weeks_with_data = {week_num for room_data in room_semester_view.values() 
                               for week_num in room_data.keys()}
        if all_weeks_with_data:
            max_weeks = max(all_weeks_with_data)

    if max_weeks == 0:
        print("Không tìm thấy dữ liệu lịch trình cho bất kỳ tuần nào. Đã dừng xuất file.")
        return

    # Tạo một workbook (tệp Excel) mới
    workbook_room = openpyxl.Workbook()
    is_first_sheet = True # Cờ để kiểm tra xem đây có phải sheet đầu tiên không (sheet mặc định)

    # Lặp qua từng tuần để tạo các sheet tương ứng
    for week_num_display in range(1, max_weeks + 1):
        if is_first_sheet:
            sheet = workbook_room.active # Sử dụng sheet mặc định
            sheet.title = f"Phong_Tuan_{week_num_display}"
            is_first_sheet = False
        else:
            # Tạo sheet mới cho các tuần tiếp theo
            sheet = workbook_room.create_sheet(title=f"Phong_Tuan_{week_num_display}")

        # --- Ghi thông tin Header của Trường Đại học ---
        current_row_for_header = 1
        # Dòng 1: Tên trung tâm bằng tiếng Việt
        sheet.cell(row=current_row_for_header, column=1, value=university_header_lines[0]).font = university_header_font
        sheet.merge_cells(start_row=current_row_for_header, start_column=1, end_row=current_row_for_header, end_column=len(excel_main_headers))
        sheet.cell(row=current_row_for_header, column=1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[current_row_for_header].height = 25
        current_row_for_header += 1
        
        # Dòng 2: Tên trung tâm bằng tiếng Anh
        sheet.cell(row=current_row_for_header, column=1, value=university_header_lines[1]).font = university_subheader_font
        sheet.merge_cells(start_row=current_row_for_header, start_column=1, end_row=current_row_for_header, end_column=len(excel_main_headers))
        sheet.cell(row=current_row_for_header, column=1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[current_row_for_header].height = 20
        current_row_for_header += 1
        
        # Dòng 3: Thông tin liên hệ
        sheet.cell(row=current_row_for_header, column=1, value=university_header_lines[2]).font = university_contact_font
        sheet.merge_cells(start_row=current_row_for_header, start_column=1, end_row=current_row_for_header, end_column=len(excel_main_headers))
        sheet.cell(row=current_row_for_header, column=1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[current_row_for_header].height = 20
        current_row_for_header += 1
        
        # Tạo một hàng trống nhỏ để phân tách
        sheet.row_dimensions[current_row_for_header].height = 10
        current_row_for_header += 1

        # --- Ghi tiêu đề của sheet (LỊCH SỬ DỤNG PHÒNG - TUẦN X) ---
        title_row = current_row_for_header
        sheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=len(excel_main_headers))
        title_cell = sheet.cell(row=title_row, column=1, value=f"LỊCH SỬ DỤNG PHÒNG - TUẦN {week_num_display}")
        title_cell.font = sheet_title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[title_row].height = 30
        
        current_excel_row = title_row + 2 # Bắt đầu hàng dữ liệu sau tiêu đề và một hàng trống
        
        # --- Ghi các header chính của bảng (Slot Thời Gian, Phòng, Thứ 2, ...) ---
        for col_idx, header_text in enumerate(excel_main_headers, start=1):
            cell = sheet.cell(row=current_excel_row, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment_center
            cell.border = cell_border
        sheet.row_dimensions[current_excel_row].height = 35 # Chiều cao hàng header
        current_excel_row += 1 # Chuyển xuống hàng tiếp theo để ghi dữ liệu

        # --- Thu thập và sắp xếp dữ liệu cho tuần hiện tại ---
        all_lessons_this_week = []
        for room_id, weekly_data in room_semester_view.items():
            if week_num_display in weekly_data:
                for lesson in weekly_data[week_num_display]:
                    lesson_copy = lesson.copy() # Tạo bản sao để tránh sửa đổi dữ liệu gốc
                    lesson_copy['room_id'] = room_id
                    
                    # Lấy tên giảng viên từ lecturer_id để hiển thị trong ô
                    lecturer_id = lesson_copy.get('lecturer_id')
                    lecturer_name = processed_data.lecturer_map.get(lecturer_id, {}).get('lecturer_name', lecturer_id)
                    lesson_copy['lecturer_name'] = lecturer_name # Thêm vào bản sao của lesson
                    
                    all_lessons_this_week.append(lesson_copy)

        # Nếu không có lịch sử dụng phòng nào cho tuần này, ghi thông báo và chuyển sang tuần tiếp theo
        if not all_lessons_this_week:
            empty_msg_cell = sheet.cell(row=current_excel_row, column=1, value="Không có phòng nào được sử dụng trong tuần này.")
            sheet.merge_cells(start_row=current_excel_row, start_column=1, end_row=current_excel_row, end_column=len(excel_main_headers))
            empty_msg_cell.alignment = Alignment(horizontal="center")
            empty_msg_cell.font = Font(italic=True, name="Times New Roman")
            
            # Áp dụng border cho các ô trong hàng thông báo rỗng
            for empty_col_idx in range(1, len(excel_main_headers) + 1):
                sheet.cell(row=current_excel_row, column=empty_col_idx).border = cell_border
            continue # Chuyển sang tuần tiếp theo

        # Sắp xếp các tiết học: Ưu tiên theo slot thời gian, sau đó theo ID phòng
        # processed_data.slot_order_map phải được đảm bảo có sẵn để sắp xếp đúng thứ tự slot
        all_lessons_this_week.sort(key=lambda x: (processed_data.slot_order_map.get(x['slot_id'], 99), x['room_id']))

        # Tạo một lưới lịch trình tạm thời để dễ dàng truy cập dữ liệu khi ghi vào Excel
        # Cấu trúc: schedule_grid[(slot_id, room_id)][day_of_week_english] = lesson_data
        schedule_grid = defaultdict(lambda: defaultdict(dict))
        for lesson in all_lessons_this_week:
            slot_id = lesson['slot_id']
            room_id = lesson['room_id']
            day = lesson['day'] # 'day' ở đây là tên ngày tiếng Anh (Mon, Tue,...)
            schedule_grid[(slot_id, room_id)][day] = lesson

        # Sắp xếp các khóa của schedule_grid để đảm bảo thứ tự hiển thị trong Excel
        sorted_schedule_keys = sorted(schedule_grid.keys(), 
                                      key=lambda x: (processed_data.slot_order_map.get(x[0], 99), x[1]))

        # --- Đổ dữ liệu lịch trình vào sheet ---
        previous_slot_id = None # Dùng để hợp nhất các ô "Slot Thời Gian"
        for slot_id, room_id in sorted_schedule_keys:
            # Ghi thông tin Slot Thời Gian và Phòng
            if slot_id != previous_slot_id:
                # Nếu slot_id thay đổi, ghi slot mới
                slot_start_time = slot_info_map.get(slot_id, {}).get('start', '')
                slot_end_time = slot_info_map.get(slot_id, {}).get('end', '')
                slot_cell = sheet.cell(row=current_excel_row, column=1, value=f"{slot_id}\n({slot_start_time}-{slot_end_time})")
                slot_cell.font = slot_time_font
                slot_cell.alignment = cell_alignment_wrapped_center
                sheet.row_dimensions[current_excel_row].height = 60 # Tăng chiều cao hàng để chứa nội dung
            
            room_cell = sheet.cell(row=current_excel_row, column=2, value=room_id)
            room_cell.alignment = cell_alignment_wrapped_center
            
            # Ghi dữ liệu cho từng ngày trong tuần
            for col_idx, day_eng in enumerate(schedule_days_eng, start=3):
                lesson_content = schedule_grid[(slot_id, room_id)].get(day_eng)
                cell = sheet.cell(row=current_excel_row, column=col_idx)
                
                if lesson_content:
                    # Lấy tên môn học và chuyển đổi loại tiết học sang tiếng Việt
                    subject_details = processed_data.subject_map.get(lesson_content['subject_id'], {})
                    subject_name = subject_details.get('name', lesson_content['subject_id'])
                    lesson_type_vi = "Lý thuyết" if lesson_content['lesson_type'] == "theory" else "Thực hành"
                    
                    # Định dạng chuỗi hiển thị nội dung tiết học
                    lesson_content_str = (
                        f"Lớp: {lesson_content['class_id']}\n"
                        f"Môn: {subject_name}\n"
                        f"GV: {lesson_content['lecturer_name']}\n" 
                        f"({lesson_type_vi})"
                    )
                    cell.value = lesson_content_str
                    # Tô màu nền dựa trên loại tiết học
                    cell.fill = theory_fill if lesson_content['lesson_type'] == "theory" else practice_fill
                else:
                    cell.value = "" # Để trống nếu không có tiết học
                
                cell.alignment = cell_alignment_wrapped_center
                cell.border = cell_border
            
            # Áp dụng border cho cột Slot Thời Gian và Phòng
            sheet.cell(row=current_excel_row, column=1).border = cell_border
            sheet.cell(row=current_excel_row, column=2).border = cell_border
            
            current_excel_row += 1 # Di chuyển xuống hàng tiếp theo
            previous_slot_id = slot_id # Cập nhật slot_id trước đó để hợp nhất

        # --- Hợp nhất các ô "Slot Thời Gian" ---
        # Bắt đầu từ hàng đầu tiên chứa dữ liệu (sau header và title)
        current_row_for_merge = title_row + 3 
        for ts_info in time_slots_info: # Lặp qua từng slot thời gian định nghĩa
            slot_id_to_merge = ts_info['slot_id']
            start_merge_row = current_row_for_merge # Hàng bắt đầu của vùng cần hợp nhất
            
            # Duyệt xuống dưới cho đến khi gặp một slot_id khác hoặc hết dữ liệu
            # Đảm bảo giá trị ô không rỗng và bắt đầu với slot_id_to_merge
            while current_row_for_merge <= sheet.max_row and \
                  sheet.cell(row=current_row_for_merge, column=1).value and \
                  str(sheet.cell(row=current_row_for_merge, column=1).value).startswith(slot_id_to_merge):
                current_row_for_merge += 1
            
            # Nếu có nhiều hơn một hàng cho cùng một slot_id, thực hiện hợp nhất
            if current_row_for_merge - start_merge_row > 1:
                sheet.merge_cells(start_row=start_merge_row, start_column=1, end_row=current_row_for_merge - 1, end_column=1)

        # --- Điều chỉnh độ rộng cột ---
        sheet.column_dimensions[get_column_letter(1)].width = 18 # Cột "Slot Thời Gian"
        sheet.column_dimensions[get_column_letter(2)].width = 15 # Cột "Phòng"
        for i in range(3, len(excel_main_headers) + 1):
            sheet.column_dimensions[get_column_letter(i)].width = 30 # Các cột ngày trong tuần

    # --- Lưu tệp Excel ---
    # Tạo thư mục đầu ra nếu chưa tồn tại
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        
    output_filename = os.path.join(output_folder, "TKB_HocKy_PhongHoc.xlsx")
    
    # Xóa tệp cũ nếu tồn tại để tránh lỗi ghi đè
    if os.path.exists(output_filename):
        try: 
            os.remove(output_filename)
            print(f"Đã xóa tệp cũ: '{output_filename}'.")
        except OSError as e:
            print(f"Lỗi khi xóa tệp Excel lịch sử dụng phòng cũ '{output_filename}': {e}. Vui lòng đóng tệp nếu đang mở.")
            return

    try:
        workbook_room.save(output_filename)
        print(f"🎉 Đã tạo thành công tệp Excel lịch sử dụng phòng (tất cả các tuần) tại: '{output_filename}'")
    except OSError as e:
        print(f"Lỗi khi lưu tệp Excel '{output_filename}': {e}. Vui lòng đảm bảo tệp không bị mở bởi ứng dụng khác.")
    except Exception as e:
        print(f"Một lỗi không mong muốn đã xảy ra khi lưu tệp Excel: {e}")
        
def export_semester_schedule_to_excel(semester_schedule_by_class: dict, processed_data, output_folder: str = "results"):
    """
    Xuất lịch trình học kỳ ra một tệp Excel duy nhất với nhiều sheets (mỗi sheet một tuần),
    bao gồm các ngày cụ thể dưới tiêu đề ngày.

    Args:
        semester_schedule_by_class (dict): Lịch trình học kỳ theo lớp.
                                           Ví dụ: {class_id: [week_0_lessons, week_1_lessons, ...]}
        processed_data (DataProcessor): Đối tượng chứa dữ liệu đã được tiền xử lý
                                        (ví dụ: thông tin lớp, giảng viên, môn học, slot thời gian).
        output_folder (str): Thư mục để lưu tệp Excel đầu ra. Mặc định là "results".
    """
    if not semester_schedule_by_class:
        print("Không có dữ liệu lịch trình học kỳ để xuất. Đã dừng thao tác.")
        return

    # --- Ánh xạ dữ liệu và Tiêu đề ---
    # Ánh xạ tên ngày tiếng Anh sang chỉ số và ngược lại
    days_of_week_index_map = {
        "Mon": 0, "Tue": 1, "Wed": 2, "Thu": 3,
        "Fri": 4, "Sat": 5, "Sun": 6
    }
    index_to_day_name_map = {
        0: "Thứ 2", 1: "Thứ 3", 2: "Thứ 4", 3: "Thứ 5",
        4: "Thứ 6", 5: "Thứ 7", 6: "Chủ Nhật"
    }
    
    # Tạo ánh xạ từ slot_id sang thông tin chi tiết của slot (start_time, end_time)
    slot_info_map = {s['slot_id']: s for s in processed_data.data.get('time_slots', [])}
    
    # Lấy danh sách các ngày trong tuần được sử dụng trong lịch trình (ví dụ: ["Mon", "Tue", ...])
    schedule_days_eng = processed_data.data.get('days_of_week', ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"])
    
    # Tạo tiêu đề cột cho các ngày trong tuần (ví dụ: "Thứ 2", "Thứ 3", ...)
    day_column_headers = [index_to_day_name_map.get(days_of_week_index_map.get(day_eng, -1), day_eng) 
                          for day_eng in schedule_days_eng]
    
    # Tiêu đề chính của bảng trong Excel
    excel_main_headers = ["Slot Thời Gian", "Lớp", "SL SV"] + day_column_headers
    
    # --- Định nghĩa Style cho Excel ---
    # Font
    header_font = Font(bold=True, color="FFFFFF", name="Times New Roman", size=11)
    sheet_title_font = Font(bold=True, size=16, name="Times New Roman")
    slot_time_font = Font(bold=True, name="Times New Roman", size=11)
    university_header_font = Font(name="Times New Roman", size=12, bold=True)
    university_subheader_font = Font(name="Times New Roman", size=11)
    university_contact_font = Font(name="Times New Roman", size=10)

    # Fill (màu nền)
    header_fill = PatternFill(start_color="4A86E8", end_color="4A86E8", fill_type="solid") # Màu xanh cho tiêu đề
    theory_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid") # Màu xanh nhạt cho tiết lý thuyết
    practice_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid") # Màu vàng nhạt cho tiết thực hành

    # Alignment (căn chỉnh)
    cell_alignment_wrapped_center = Alignment(wrap_text=True, horizontal='center', vertical='center')
    header_alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Border (đường viền)
    thin_border_side = Side(style='thin', color="000000")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Thông tin header của trường đại học
    university_header_lines = [
        "TRUNG TÂM CÔNG NGHỆ PHẦN MỀM ĐẠI HỌC CẦN THƠ",
        "CANTHO UNIVERSITY SOFTWARE CENTER",
        "Khu III, Đại học Cần Thơ - 01 Lý Tự Trọng, TP. Cần Thơ - Tel: 0292.3731072 & Fax: 0292.3731071 - Email: cusc@ctu.edu.vn"
    ]

    # --- Xác định số tuần tối đa của học kỳ ---
    max_weeks = 0
    # Ưu tiên lấy từ duration_weeks trong semester_map của processed_data nếu có
    if processed_data and hasattr(processed_data, 'semester_map'):
        for sem_info in processed_data.semester_map.values():
            max_weeks = max(max_weeks, sem_info.get('duration_weeks', 0))
    
    # Nếu không có thông tin duration_weeks hoặc dữ liệu rỗng, kiểm tra từ schedule_by_class
    if max_weeks == 0 and semester_schedule_by_class:
        for schedules_list in semester_schedule_by_class.values():
            max_weeks = max(max_weeks, len(schedules_list))
    
    if max_weeks == 0:
        print("Không thể xác định thời lượng học kỳ (số tuần). Hủy xuất file.")
        return
        
    # Tạo một workbook (tệp Excel) mới
    workbook = openpyxl.Workbook()
    is_first_sheet = True # Cờ để kiểm tra xem đây có phải sheet đầu tiên không (sheet mặc định)

    # Lặp qua từng tuần để tạo các sheet tương ứng
    for week_idx in range(max_weeks):
        week_num_display = week_idx + 1 # Tuần hiển thị bắt đầu từ 1
        
        if is_first_sheet:
            sheet = workbook.active # Sử dụng sheet mặc định
            sheet.title = f"Tuan_{week_num_display}"
            is_first_sheet = False
        else:
            # Tạo sheet mới cho các tuần tiếp theo
            sheet = workbook.create_sheet(title=f"Tuan_{week_num_display}")

        # --- Ghi thông tin Header của Trường Đại học ---
        current_row_for_header = 1
        # Dòng 1: Tên trung tâm bằng tiếng Việt
        sheet.cell(row=current_row_for_header, column=1, value=university_header_lines[0]).font = university_header_font
        sheet.merge_cells(start_row=current_row_for_header, start_column=1, end_row=current_row_for_header, end_column=len(excel_main_headers))
        sheet.cell(row=current_row_for_header, column=1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[current_row_for_header].height = 25
        current_row_for_header += 1
        
        # Dòng 2: Tên trung tâm bằng tiếng Anh
        sheet.cell(row=current_row_for_header, column=1, value=university_header_lines[1]).font = university_subheader_font
        sheet.merge_cells(start_row=current_row_for_header, start_column=1, end_row=current_row_for_header, end_column=len(excel_main_headers))
        sheet.cell(row=current_row_for_header, column=1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[current_row_for_header].height = 20
        current_row_for_header += 1
        
        # Dòng 3: Thông tin liên hệ
        sheet.cell(row=current_row_for_header, column=1, value=university_header_lines[2]).font = university_contact_font
        sheet.merge_cells(start_row=current_row_for_header, start_column=1, end_row=current_row_for_header, end_column=len(excel_main_headers))
        sheet.cell(row=current_row_for_header, column=1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[current_row_for_header].height = 20
        current_row_for_header += 1
        
        # Tạo một hàng trống nhỏ để phân tách
        sheet.row_dimensions[current_row_for_header].height = 10
        current_row_for_header += 1

        # --- Ghi tiêu đề của sheet (THỜI KHÓA BIỂU TUẦN X) ---
        sheet.merge_cells(start_row=current_row_for_header, start_column=1, end_row=current_row_for_header, end_column=len(excel_main_headers))
        title_cell = sheet.cell(row=current_row_for_header, column=1, value=f"THỜI KHÓA BIỂU TUẦN {week_num_display}")
        title_cell.font = sheet_title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[current_row_for_header].height = 30
        
        # --- Ghi các header chính của bảng (Slot Thời Gian, Lớp, SL SV, Thứ 2, ...) ---
        current_excel_row = current_row_for_header + 2 
        for col_idx, header_text in enumerate(excel_main_headers, start=1):
            cell = sheet.cell(row=current_excel_row, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment_center
            cell.border = cell_border
        sheet.row_dimensions[current_excel_row].height = 35 # Chiều cao hàng header
        
        # --- THÊM HÀNG CHỨA NGÀY CỤ THỂ DƯỚI HEADER NGÀY ---
        current_excel_row += 1
        
        # Tìm một lesson bất kỳ trong tuần này để lấy ngày bắt đầu tính toán
        first_lesson_in_week = None
        for schedules_list in semester_schedule_by_class.values():
            if week_idx < len(schedules_list) and schedules_list[week_idx]:
                first_lesson_in_week = schedules_list[week_idx][0]
                break

        # Nếu tìm thấy tiết học đầu tiên và có thông tin ngày
        if first_lesson_in_week and 'date' in first_lesson_in_week:
            # Chuyển đổi ngày từ string sang đối tượng datetime
            base_date = datetime.strptime(first_lesson_in_week['date'], '%Y-%m-%d')
            # Tìm ngày đầu tuần (Thứ 2) của tuần chứa base_date
            # weekday() trả về 0 cho Thứ 2, 6 cho Chủ Nhật
            start_date_of_week = base_date - timedelta(days=base_date.weekday()) 

            # Ghi ô "Ngày" cho 3 cột đầu tiên
            date_label_cell = sheet.cell(row=current_excel_row, column=1, value="Ngày")
            date_label_cell.font = header_font
            date_label_cell.fill = header_fill
            date_label_cell.alignment = header_alignment_center
            sheet.merge_cells(start_row=current_excel_row, start_column=1, end_row=current_excel_row, end_column=3)
            
            # Ghi các ngày cụ thể cho từng cột ngày trong tuần
            for col_idx, day_eng in enumerate(schedule_days_eng, start=4):
                # Tính toán ngày thực tế cho từng ngày trong tuần
                # Lấy index của ngày tiếng Anh hiện tại và cộng vào start_date_of_week
                day_offset = days_of_week_index_map.get(day_eng, 0)
                date_for_day = start_date_of_week + timedelta(days=day_offset)
                
                cell = sheet.cell(row=current_excel_row, column=col_idx, value=date_for_day.strftime('%d/%m'))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment_center
                cell.border = cell_border
        
        sheet.row_dimensions[current_excel_row].height = 25 # Chiều cao hàng ngày
        # --- KẾT THÚC PHẦN THÊM HÀNG CHỨA NGÀY CỤ THỂ ---

        current_excel_row += 1 # Chuyển xuống hàng tiếp theo để ghi dữ liệu lịch trình

        # --- Thu thập và sắp xếp dữ liệu lịch trình cho tuần hiện tại ---
        all_lessons_this_week = []
        for class_id_key, weekly_schedules_val in semester_schedule_by_class.items():
            if week_idx < len(weekly_schedules_val) and weekly_schedules_val[week_idx]:
                # Thêm class_id vào mỗi lesson để dễ dàng nhóm và sắp xếp
                for lesson in weekly_schedules_val[week_idx]:
                    lesson_copy = lesson.copy() # Tạo bản sao để tránh sửa đổi dữ liệu gốc
                    lesson_copy['class_id'] = class_id_key # Đảm bảo class_id đúng
                    all_lessons_this_week.append(lesson_copy)

        # Nếu không có lịch học nào trong tuần này, ghi thông báo và chuyển sang tuần tiếp theo
        if not all_lessons_this_week:
            empty_msg_cell = sheet.cell(row=current_excel_row, column=1, value="Không có lịch học trong tuần này.")
            empty_msg_cell.font = Font(italic=True, name="Times New Roman", size=11)
            sheet.merge_cells(start_row=current_excel_row, start_column=1, end_row=current_excel_row, end_column=len(excel_main_headers))
            empty_msg_cell.alignment = Alignment(horizontal="center")
            
            # Áp dụng border cho các ô trong hàng thông báo rỗng
            for empty_col_idx in range(1, len(excel_main_headers) + 1):
                sheet.cell(row=current_excel_row, column=empty_col_idx).border = cell_border
            continue # Chuyển sang tuần tiếp theo
        
        # Sắp xếp các tiết học: Ưu tiên theo slot thời gian, sau đó theo ID lớp
        # Đảm bảo processed_data.slot_order_map tồn tại và có dữ liệu
        if not hasattr(processed_data, 'slot_order_map') or not processed_data.slot_order_map:
            print("Cảnh báo: processed_data.slot_order_map không khả dụng hoặc rỗng. Sắp xếp theo slot_id string.")
            all_lessons_this_week.sort(key=lambda x: (x['slot_id'], x['class_id']))
        else:
            all_lessons_this_week.sort(key=lambda x: (processed_data.slot_order_map.get(x['slot_id'], 9999), x['class_id'])) # Sử dụng giá trị mặc định lớn nếu slot_id không có trong map
        
        # Tạo một lưới lịch trình tạm thời để dễ dàng truy cập dữ liệu khi ghi vào Excel
        # Cấu trúc: schedule_grid[(slot_id, class_id)][day_of_week_english] = lesson_data
        schedule_grid = defaultdict(lambda: defaultdict(dict))
        for lesson in all_lessons_this_week:
            slot_id = lesson['slot_id']
            class_id = lesson['class_id']
            day = lesson['day'] # 'day' ở đây là tên ngày tiếng Anh (Mon, Tue,...)
            schedule_grid[(slot_id, class_id)][day] = lesson

        # Sắp xếp các khóa của schedule_grid để đảm bảo thứ tự hiển thị trong Excel
        if not hasattr(processed_data, 'slot_order_map') or not processed_data.slot_order_map:
             sorted_schedule_keys = sorted(schedule_grid.keys(), key=lambda x: (x[0], x[1]))
        else:
            sorted_schedule_keys = sorted(schedule_grid.keys(), key=lambda x: (processed_data.slot_order_map.get(x[0], 9999), x[1]))

        # --- Đổ dữ liệu lịch trình vào sheet ---
        # `previous_slot_id` không cần thiết ở đây vì ô "Slot Thời Gian" sẽ được hợp nhất sau.
        for slot_id, class_id in sorted_schedule_keys:
            class_info = processed_data.class_map.get(class_id, {})
            class_size = class_info.get('size', 'N/A') # Lấy sĩ số lớp
            
            # Ghi thông tin Slot Thời Gian, Lớp, Sĩ số
            # Lấy thời gian bắt đầu/kết thúc từ slot_info_map
            slot_start_time = slot_info_map.get(slot_id, {}).get('start', '')
            slot_end_time = slot_info_map.get(slot_id, {}).get('end', '')

            slot_cell = sheet.cell(row=current_excel_row, column=1, value=f"{slot_id}\n({slot_start_time}-{slot_end_time})")
            slot_cell.font = slot_time_font
            slot_cell.alignment = cell_alignment_wrapped_center
            sheet.row_dimensions[current_excel_row].height = 60 # Chiều cao hàng để chứa nội dung
            
            class_cell = sheet.cell(row=current_excel_row, column=2, value=class_id)
            slsv_cell = sheet.cell(row=current_excel_row, column=3, value=class_size)
            
            class_cell.alignment = cell_alignment_wrapped_center
            slsv_cell.alignment = cell_alignment_wrapped_center
            
            # Ghi dữ liệu cho từng ngày trong tuần
            for col_idx, day_eng in enumerate(schedule_days_eng, start=4):
                lesson_content = schedule_grid[(slot_id, class_id)].get(day_eng)
                cell = sheet.cell(row=current_excel_row, column=col_idx)
                
                if lesson_content:
                    # Lấy tên môn học và giảng viên, chuyển đổi loại tiết học sang tiếng Việt
                    subject_details = processed_data.subject_map.get(lesson_content['subject_id'], {})
                    lecturer_details = processed_data.lecturer_map.get(lesson_content['lecturer_id'], {})
                    subject_name = subject_details.get('name', lesson_content['subject_id'])
                    lecturer_name = lecturer_details.get('lecturer_name', lesson_content['lecturer_id'])
                    lesson_type_vi = "Lý thuyết" if lesson_content['lesson_type'] == "theory" else "Thực hành"
                    
                    # Định dạng chuỗi hiển thị nội dung tiết học
                    lesson_content_str = (
                        f"{subject_name}\n"
                        f"({lesson_type_vi})\n"
                        f"Phòng: {lesson_content['room_id']}\n"
                        f"GV: {lecturer_name}"
                    )
                    cell.value = lesson_content_str
                    # Tô màu nền dựa trên loại tiết học
                    cell.fill = theory_fill if lesson_content['lesson_type'] == "theory" else practice_fill
                else:
                    cell.value = "" # Để trống nếu không có tiết học
                
                cell.alignment = cell_alignment_wrapped_center
                cell.border = cell_border
            
            # Áp dụng border cho 3 cột đầu tiên
            sheet.cell(row=current_excel_row, column=1).border = cell_border
            sheet.cell(row=current_excel_row, column=2).border = cell_border
            sheet.cell(row=current_excel_row, column=3).border = cell_border
            
            current_excel_row += 1 # Di chuyển xuống hàng tiếp theo

        # --- Hợp nhất các ô "Slot Thời Gian" ---
        # Bắt đầu từ hàng đầu tiên chứa dữ liệu (sau các header)
        # current_row_for_merge = title_row + 3 nếu không có hàng ngày
        # current_row_for_merge = title_row + 4 nếu có hàng ngày
        current_row_for_merge_start = current_row_for_header + 4 # Sau title, main headers, và date row
        
        for slot_info in processed_data.data.get('time_slots', []):
            slot_id_to_merge = slot_info['slot_id']
            start_merge_row = current_row_for_merge_start
            
            # Duyệt xuống dưới cho đến khi gặp một slot_id khác hoặc hết dữ liệu
            # Đảm bảo giá trị ô không rỗng và bắt đầu với slot_id_to_merge
            while start_merge_row <= sheet.max_row and \
                  sheet.cell(row=start_merge_row, column=1).value and \
                  str(sheet.cell(row=start_merge_row, column=1).value).startswith(f"{slot_id_to_merge}\n"):
                start_merge_row += 1
            
            # Nếu có nhiều hơn một hàng cho cùng một slot_id, thực hiện hợp nhất
            if start_merge_row - current_row_for_merge_start > 1:
                sheet.merge_cells(start_row=current_row_for_merge_start, start_column=1, end_row=start_merge_row - 1, end_column=1)
            
            current_row_for_merge_start = start_merge_row # Cập nhật hàng bắt đầu cho slot tiếp theo

        # --- Điều chỉnh độ rộng cột ---
        sheet.column_dimensions[get_column_letter(1)].width = 20 # Cột "Slot Thời Gian"
        sheet.column_dimensions[get_column_letter(2)].width = 15 # Cột "Lớp"
        sheet.column_dimensions[get_column_letter(3)].width = 8  # Cột "SL SV"
        for i in range(4, len(excel_main_headers) + 1):
            sheet.column_dimensions[get_column_letter(i)].width = 28 # Các cột ngày trong tuần

    # --- Lưu tệp Excel ---
    # Tạo thư mục đầu ra nếu chưa tồn tại
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        
    output_filename = os.path.join(output_folder, "TKB_Semester_Schedule.xlsx")
    
    # Xóa tệp cũ nếu tồn tại để tránh lỗi ghi đè và lỗi PermissionError
    if os.path.exists(output_filename):
        try: 
            os.remove(output_filename)
            print(f"Đã xóa tệp cũ: '{output_filename}'.")
        except OSError as e:
            print(f"Lỗi khi xóa tệp Excel lịch học kỳ cũ '{output_filename}': {e}. Vui lòng đóng tệp nếu đang mở.")
            return

    try:
        workbook.save(output_filename)
        print(f"🎉 Đã tạo thành công tệp Excel lịch học kỳ (tất cả các tuần) tại: '{output_filename}'")
    except OSError as e:
        print(f"Lỗi khi lưu tệp Excel '{output_filename}': {e}. Vui lòng đảm bảo tệp không bị mở bởi ứng dụng khác.")
    except Exception as e:
        print(f"Một lỗi không mong muốn đã xảy ra khi lưu tệp Excel: {e}")