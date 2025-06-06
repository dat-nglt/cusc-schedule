import openpyxl, os
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict

def export_semester_schedule_to_excel(semester_schedule_by_class, processed_data, output_folder="results"):
    """
    Exports the semester schedule to multiple Excel files, one for each week,
    with days of the week as columns and enhanced formatting.

    Args:
        semester_schedule_by_class (dict): The semester schedule, 
            e.g., {class_id: [week_0_lessons, week_1_lessons, ...]}
        processed_data (DataProcessor): The processed data object.
        output_folder (str): Folder to save the Excel files.
    """
    if not semester_schedule_by_class:
        print("No semester schedule data to export.")
        return

    days_of_week_map = {
        "Mon": "Thứ 2", "Tue": "Thứ 3", "Wed": "Thứ 4",
        "Thu": "Thứ 5", "Fri": "Thứ 6", "Sat": "Thứ 7", "Sun": "Chủ Nhật"
    }
    schedule_days_eng = processed_data.data.get('days_of_week', ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"])
    # Tạo tiêu đề cột cho các Thứ trong tuần
    day_column_headers = [days_of_week_map.get(day_eng, day_eng) for day_eng in schedule_days_eng]
    
    time_slots_info = processed_data.data['time_slots']
    
    # Tiêu đề chính cho bảng Excel
    excel_main_headers = ["Slot Thời Gian", "Lớp", "SL SV"] + day_column_headers

    # Định nghĩa các Styles để tái sử dụng
    header_font = Font(bold=True, color="FFFFFF", name="Times New Roman", size=11)
    header_fill = PatternFill(start_color="4A86E8", end_color="4A86E8", fill_type="solid") # Màu xanh dương nhạt hơn
    sheet_title_font = Font(bold=True, size=16, name="Times New Roman")
    cell_alignment_wrapped_center = Alignment(wrap_text=True, horizontal='center', vertical='center')
    header_alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border_side = Side(style='thin', color="000000")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    slot_time_font = Font(bold=True, name="Times New Roman", size=11)
    
    # Define colors for lesson types
    theory_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Light blue for theory
    practice_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # Light orange for practice

    # University header information
    university_header = [
        "TRUNG TÂM CÔNG NGHỆ PHẦN MỀM ĐẠI HỌC CẦN THƠ",
        "CANTHO UNIVERSITY SOFTWARE CENTER",
        "Khu III, Đại học Cần Thơ - 01 Lý Tự Trọng, TP. Cần Thơ - Tel: 0292.3731072 & Fax: 0292.3731071 - Email: cusc@ctu.edu.vn"
    ]
    university_header_font = Font(name="Times New Roman", size=12, bold=True)
    university_subheader_font = Font(name="Times New Roman", size=11)
    university_contact_font = Font(name="Times New Roman", size=10)

    # Xác định số tuần tối đa cần tạo file
    max_weeks = 0
    if semester_schedule_by_class: 
        if any(s for s_list in semester_schedule_by_class.values() for s in s_list): # Check if any lesson exists
            max_weeks = max(len(schedules) for schedules in semester_schedule_by_class.values() if schedules)
        if max_weeks == 0 and processed_data.class_map:
             max_weeks = max(cls.get('program_duration_weeks', 0) for cls in processed_data.class_map.values())

    for week_idx in range(max_weeks):
        week_num_display = week_idx + 1
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"Tuan_{week_num_display}"

        # --- Add University Header ---
        # Row 1: Vietnamese name
        sheet.cell(row=1, column=1, value=university_header[0]).font = university_header_font
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(excel_main_headers))
        sheet.row_dimensions[1].height = 25
        
        # Row 2: English name
        sheet.cell(row=2, column=1, value=university_header[1]).font = university_subheader_font
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(excel_main_headers))
        sheet.row_dimensions[2].height = 20
        
        # Row 3: Contact information
        sheet.cell(row=3, column=1, value=university_header[2]).font = university_contact_font
        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(excel_main_headers))
        sheet.row_dimensions[3].height = 20
        
        # Empty row for spacing
        sheet.row_dimensions[4].height = 10

        # --- Bắt đầu định dạng ---

        # 1. Tiêu đề chính của Sheet (moved to row 5)
        sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=len(excel_main_headers))
        title_cell = sheet.cell(row=5, column=1, value=f"THỜI KHÓA BIỂU TUẦN {week_num_display}")
        title_cell.font = sheet_title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[5].height = 30
        
        current_excel_row = 7  # Adjusted starting row due to added headers

        # 2. Ghi dòng tiêu đề chính của bảng (Slot Thời Gian, Lớp, SL SV, Thứ 2, Thứ 3, ...)
        for col_idx, header_text in enumerate(excel_main_headers, start=1):
            cell = sheet.cell(row=current_excel_row, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment_center
            cell.border = cell_border
        sheet.row_dimensions[current_excel_row].height = 35 
        current_excel_row += 1

        # Thu thập tất cả các tiết học của tuần này để xử lý
        all_lessons_this_week = []
        for class_id_key, weekly_schedules_val in semester_schedule_by_class.items():
            if week_idx < len(weekly_schedules_val) and weekly_schedules_val[week_idx]:
                all_lessons_this_week.extend(weekly_schedules_val[week_idx])
        
        if not all_lessons_this_week: 
            empty_msg_cell = sheet.cell(row=current_excel_row, column=1, value="Không có lịch học trong tuần này.")
            empty_msg_cell.font = Font(italic=True, name="Times New Roman", size=11)
            sheet.merge_cells(start_row=current_excel_row, start_column=1, end_row=current_excel_row, end_column=len(excel_main_headers))
            empty_msg_cell.alignment = Alignment(horizontal="center")
            workbook.save(f"{output_folder}/TKB_Tuan_{week_num_display}_TheoThu.xlsx")
            print(f"Generated Excel for Week {week_num_display} (empty schedule) with new day-column format.")
            continue

        # 3. Lặp qua từng Slot Thời Gian để điền dữ liệu
        row_start_for_current_slot_merge = current_excel_row 
        for ts_info in time_slots_info:
            slot_id_current = ts_info['slot_id']
            slot_display_text = f"{ts_info['slot_id']}\n({ts_info['start']}-{ts_info['end']})"
            
            # Lấy danh sách các lớp có lịch vào Slot Thời Gian này (đã sắp xếp)
            classes_in_this_slot = sorted(list(set(
                l['class_id'] for l in all_lessons_this_week if l['slot_id'] == slot_id_current
            )))

            if not classes_in_this_slot: 
                # Option: Ghi dòng cho slot trống hoặc bỏ qua
                # Ghi dòng cho slot trống:
                slot_cell_empty = sheet.cell(row=current_excel_row, column=1, value=slot_display_text)
                slot_cell_empty.alignment = cell_alignment_wrapped_center
                slot_cell_empty.border = cell_border
                slot_cell_empty.font = slot_time_font
                # Áp dụng border cho các ô còn lại trong dòng slot trống
                for empty_col_idx in range(2, len(excel_main_headers) + 1):
                    sheet.cell(row=current_excel_row, column=empty_col_idx).border = cell_border
                sheet.row_dimensions[current_excel_row].height = 45
                current_excel_row += 1
                row_start_for_current_slot_merge = current_excel_row # Cập nhật cho slot tiếp theo
                continue

            for i, class_id in enumerate(classes_in_this_slot):
                class_info = processed_data.class_map.get(class_id, {})
                class_size = class_info.get('size', 'N/A')
                
                # Cột "Slot Thời Gian"
                if i == 0: 
                    slot_cell = sheet.cell(row=current_excel_row, column=1, value=slot_display_text)
                    slot_cell.font = slot_time_font # In đậm Slot
                else: 
                    slot_cell = sheet.cell(row=current_excel_row, column=1, value="") 
                slot_cell.alignment = cell_alignment_wrapped_center
                slot_cell.border = cell_border


                # Cột "Lớp"
                class_cell = sheet.cell(row=current_excel_row, column=2, value=class_id)
                class_cell.alignment = cell_alignment_wrapped_center
                class_cell.border = cell_border
                class_cell.font = Font(name="Times New Roman", size=11)

                # Cột "SL SV"
                slsv_cell = sheet.cell(row=current_excel_row, column=3, value=class_size)
                slsv_cell.alignment = cell_alignment_wrapped_center
                slsv_cell.border = cell_border
                slsv_cell.font = Font(name="Times New Roman", size=11)

                # Các cột Thứ trong tuần
                current_data_col = 4 
                for day_eng in schedule_days_eng:
                    lesson_content_str = ""
                    found_lesson_for_day_slot_class = None
                    
                    for lesson in all_lessons_this_week:
                        if lesson['class_id'] == class_id and \
                           lesson['slot_id'] == slot_id_current and \
                           lesson['day'] == day_eng:
                            found_lesson_for_day_slot_class = lesson
                            break 
                    
                    if found_lesson_for_day_slot_class:
                        subject_details = processed_data.subject_map.get(found_lesson_for_day_slot_class['subject_id'], {})
                        subject_name = subject_details.get('name', found_lesson_for_day_slot_class['subject_id'])
                        
                        lesson_type_vi = "Lý thuyết" if found_lesson_for_day_slot_class['lesson_type'] == "theory" else "Thực hành"
                        lesson_content_str = (
                            f"{subject_name}\n"
                            f"({lesson_type_vi})\n"
                            f"Phòng: {found_lesson_for_day_slot_class['room_id']}\n"
                            f"GV: {found_lesson_for_day_slot_class['lecturer_name']}"
                        )

                        # Apply color based on lesson type
                        lesson_fill = theory_fill if found_lesson_for_day_slot_class['lesson_type'] == "theory" else practice_fill
                    
                    data_cell = sheet.cell(row=current_excel_row, column=current_data_col, value=lesson_content_str)
                    data_cell.alignment = cell_alignment_wrapped_center
                    data_cell.border = cell_border
                    data_cell.font = Font(name="Times New Roman", size=10)
                    if found_lesson_for_day_slot_class:
                        data_cell.fill = lesson_fill
                    current_data_col += 1
                
                sheet.row_dimensions[current_excel_row].height = 60 
                current_excel_row += 1
            
            if classes_in_this_slot:
                num_classes_in_slot = len(classes_in_this_slot)
                if num_classes_in_slot > 1:
                    sheet.merge_cells(start_row=row_start_for_current_slot_merge, start_column=1,
                                      end_row=row_start_for_current_slot_merge + num_classes_in_slot - 1, end_column=1)
                row_start_for_current_slot_merge = current_excel_row 

        # 4. Điều chỉnh độ rộng cột
        sheet.column_dimensions[get_column_letter(1)].width = 20 # Slot Thời Gian
        sheet.column_dimensions[get_column_letter(2)].width = 15 # Lớp
        sheet.column_dimensions[get_column_letter(3)].width = 8  # SL SV
        for i in range(4, len(excel_main_headers) + 1): # Các cột Thứ
            sheet.column_dimensions[get_column_letter(i)].width = 28 
        
        output_filename = f"{output_folder}/TKB_Tuan_{week_num_display}_TheoThu.xlsx"

        # Xóa file cũ nếu tồn tại
        if os.path.exists(output_filename):
            os.remove(output_filename)

        # Lưu file mới
        workbook.save(output_filename)