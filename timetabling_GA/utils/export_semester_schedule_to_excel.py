import os
from datetime import datetime, timedelta
from collections import defaultdict
from typing import Dict, Any, Optional
from pprint import pprint
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def export_semester_schedule_to_excel(
    semester_schedule_json: Dict[str, Any], 
    output_folder: str = "results",
    semester_id: Optional[str] = None
) -> str:
    """
    Exports a semester timetable from JSON data to an Excel file, with one sheet per week.

    This function creates an Excel workbook and a separate sheet for each academic week.
    It populates the cells with detailed lesson information such as subject name,
    lesson type, room, and lecturer. Cells are formatted with colors and borders for
    improved readability.

    Args:
        semester_schedule_json (Dict[str, Any]): Timetable data for a single semester in JSON format.
        output_folder (str): The path to the directory where the Excel file will be saved.
        semester_id (Optional[str]): The ID of the semester, used for unique file naming.
    
    Returns:
        str: The path to the saved Excel file.
    """
    if not semester_schedule_json:
        print("DEBUG: Dữ liệu đầu vào JSON trống. Dừng hoạt động.")
        return ""

    print("DEBUG: Bắt đầu quá trình xuất file Excel...")
    
    # --- Configuration for mappings and styles ---
    days_of_week_index_map = {
        "Mon": 0, "Tue": 1, "Wed": 2, "Thu": 3,
        "Fri": 4, "Sat": 5, "Sun": 6
    }
    index_to_day_name_map = {
        0: "Monday", 1: "Tuesday", 2: "Wednesday", 3: "Thursday",
        4: "Friday", 5: "Saturday", 6: "Sunday"
    }
    
    # Find the start and end dates of the semester
    all_dates = []
    print("DEBUG: Bắt đầu tìm kiếm ngày tháng trong dữ liệu...")
    for class_schedules in semester_schedule_json.values():
        for week_schedule in class_schedules:
            for lesson in week_schedule:
                if 'date' in lesson:
                    try:
                        all_dates.append(datetime.strptime(lesson['date'], '%Y-%m-%d'))
                    except ValueError as e:
                        print(f"DEBUG ERROR: Định dạng ngày không hợp lệ '{lesson.get('date')}'. Lỗi: {e}")
                else:
                    print(f"DEBUG ERROR: Không tìm thấy khóa 'date' trong một bài học. Dữ liệu: {lesson}")
    
    if not all_dates:
        print("DEBUG: Không tìm thấy ngày bài học nào trong dữ liệu. Dừng hoạt động.")
        return ""

    start_date = min(all_dates)
    last_lesson_date = max(all_dates)
    
    print(f"DEBUG: Ngày bắt đầu học kỳ được xác định là {start_date.strftime('%Y-%m-%d')}")
    print(f"DEBUG: Ngày kết thúc học kỳ được xác định là {last_lesson_date.strftime('%Y-%m-%d')}")
    
    days_to_monday = days_of_week_index_map.get(start_date.strftime('%a'), 0)
    week_start_date = start_date - timedelta(days=days_to_monday)
    
    weeks = []
    current_week_start = week_start_date
    while current_week_start <= last_lesson_date:
        week_end_date = current_week_start + timedelta(days=6)
        weeks.append((current_week_start, week_end_date))
        current_week_start += timedelta(days=7)
    
    print(f"DEBUG: Đã xác định được {len(weeks)} tuần học.")

    # --- Create Excel workbook ---
    workbook = openpyxl.Workbook()
    if len(workbook.worksheets) > 0:
        workbook.remove(workbook.worksheets[0])
    print("DEBUG: Workbook mới đã được tạo.")

    # --- Define cell styles ---
    header_font = Font(bold=True, color="FFFFFF", name="Times New Roman", size=11)
    sheet_title_font = Font(bold=True, size=16, name="Times New Roman")
    slot_time_font = Font(bold=True, name="Times New Roman", size=11)
    university_header_font = Font(name="Times New Roman", size=12, bold=True)
    university_subheader_font = Font(name="Times New Roman", size=11)
    university_contact_font = Font(name="Times New Roman", size=10)

    header_fill = PatternFill(start_color="4A86E8", end_color="4A86E8", fill_type="solid")
    theory_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    practice_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
    
    cell_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    thin_border = Side(style='thin', color="000000")
    cell_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    # --- Create sheets for each week ---
    for week_num, (week_start, week_end) in enumerate(weeks, 1):
        sheet = workbook.create_sheet(title=f"Week_{week_num}")
        print(f"DEBUG: Đang tạo sheet cho Tuần {week_num}: {week_start.strftime('%d/%m/%Y')} - {week_end.strftime('%d/%m/%Y')}")
        
        # --- University header and sheet title ---
        current_row = 1
        university_header = [
            "CANTHO UNIVERSITY SOFTWARE CENTER",
            "TRUNG TÂM PHẦN MỀM ĐẠI HỌC CẦN THƠ",
            "Khu 3, Đại học Cần Thơ - 01 Lý Tự Trọng, TP Cần Thơ - Tel: 0292.3731072 & Fax: 0292.3731071 - Email: cusc@ctu.edu.vn"
        ]
        
        for i, line in enumerate(university_header):
            sheet.cell(row=current_row, column=1, value=line)
            sheet.cell(row=current_row, column=1).font = (
                university_header_font if i == 0 else 
                university_subheader_font if i == 1 else 
                university_contact_font
            )
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            sheet.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1
        
        current_row += 1
        
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
        title_cell = sheet.cell(row=current_row, column=1, value=f"WEEK {week_num} TIMETABLE ({week_start.strftime('%d/%m/%Y')} - {week_end.strftime('%d/%m/%Y')})")
        title_cell.font = sheet_title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 2
        
        # --- Column headers (Time Slot, Class, Student Count, and days of the week) ---
        headers = ["Time Slot", "Class", "Students"]
        current_date = week_start
        for i in range(7):
            day_display = f"{index_to_day_name_map[i]}\n{current_date.strftime('%d/%m')}"
            headers.append(day_display)
            current_date += timedelta(days=1)

        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = cell_alignment
            cell.border = cell_border
        
        sheet.row_dimensions[current_row].height = 40
        current_row += 1
        
        # --- Group data by slot and class for sorting and display ---
        schedule_grid = defaultdict(lambda: defaultdict(dict))
        total_lessons_in_week = 0
        for class_id, class_schedules in semester_schedule_json.items():
            for week_schedule in class_schedules:
                for lesson in week_schedule:
                    if 'date' in lesson:
                        lesson_date = datetime.strptime(lesson['date'], '%Y-%m-%d')
                        if week_start <= lesson_date <= week_end:
                            schedule_grid[(lesson['slot_id'], class_id)][lesson['day']] = lesson
                            total_lessons_in_week += 1
        
        print(f"DEBUG: Số lượng bài học được tìm thấy cho tuần {week_num}: {len(schedule_grid)}")

        # --- Populate the timetable data into the sheet ---
        sorted_keys = sorted(schedule_grid.keys())
        
        if not sorted_keys:
            print(f"DEBUG: Tuần {week_num} không có bài học nào. Bỏ qua việc điền dữ liệu.")
            # Có thể thêm logic để thông báo trên sheet rằng không có lịch học
            current_row += 1
            empty_cell = sheet.cell(row=current_row, column=1, value="Không có lịch học trong tuần này.")
            empty_cell.alignment = Alignment(horizontal="center")
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            
        else:
            for slot_id, class_id in sorted_keys:
                day_lessons = schedule_grid[(slot_id, class_id)]
                
                class_size = next((l.get('size', 'N/A') for l in day_lessons.values()), 'N/A')
                
                sheet.cell(row=current_row, column=1, value=slot_id).font = slot_time_font
                sheet.cell(row=current_row, column=2, value=class_id)
                sheet.cell(row=current_row, column=3, value=class_size)
                
                current_date = week_start
                for col_idx in range(4, 11):
                    day_name = current_date.strftime('%a')
                    # Điều kiện này kiểm tra xem ngày bài học có lớn hơn ngày bắt đầu học kỳ không
                    if day_name in day_lessons and current_date >= start_date:
                        lesson = day_lessons[day_name]
                        # Kiểm tra xem khóa 'lesson_type' có tồn tại không
                        lesson_type = lesson.get('lesson_type', 'N/A')
                        lesson_type_display = "Theory" if lesson_type == "theory" else "Practice" if lesson_type == "practice" else "N/A"
                        
                        # Sử dụng .get() để tránh KeyError nếu khóa không tồn tại
                        content = f"{lesson.get('subject_id', 'N/A')}\n({lesson_type_display})\nRoom: {lesson.get('room_id', 'N/A')}\nLecturer: {lesson.get('lecturer_id', 'N/A')}"
                        
                        cell = sheet.cell(row=current_row, column=col_idx, value=content)
                        # Gán màu dựa trên lesson_type
                        if lesson_type == "theory":
                            cell.fill = theory_fill
                        elif lesson_type == "practice":
                            cell.fill = practice_fill
                    else:
                        cell = sheet.cell(row=current_row, column=col_idx, value="")

                    cell.alignment = cell_alignment
                    cell.border = cell_border
                    current_date += timedelta(days=1)
                
                for col in [1, 2, 3]:
                    sheet.cell(row=current_row, column=col).alignment = cell_alignment
                    sheet.cell(row=current_row, column=col).border = cell_border
                
                sheet.row_dimensions[current_row].height = 60
                current_row += 1

            merge_slot_cells(sheet, start_row=current_row - len(sorted_keys))
        
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 8
        for col_idx in range(4, 11):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 25
    
    # --- Save the Excel file ---
    os.makedirs(output_folder, exist_ok=True)
    current_dir = os.path.dirname(os.path.abspath(__file__))
    repo_root = os.path.dirname(os.path.dirname(current_dir))
    be_results_dir = os.path.join(repo_root, 'be', 'results')
    
    if semester_id:
        output_path = os.path.join(output_folder, f"LichHocKy_{semester_id}.xlsx")
        output_be_path = os.path.join(be_results_dir, f"LichHocKy_{semester_id}.xlsx")
    else:
        output_path = os.path.join(output_folder, "Semester_Timetable.xlsx")
        output_be_path = os.path.join(be_results_dir, "Semester_Timetable.xlsx")
    
    try:
        if os.path.exists(output_path):
            os.remove(output_path)
            print(f"DEBUG: Đã xóa file cũ tại: {output_path}")
        workbook.save(output_path)
        print(f"\nDEBUG: Đã xuất file Excel thành công tới: {output_path}")
        
        if os.path.exists(output_be_path):
            os.remove(output_be_path)
            print(f"DEBUG: Đã xóa file cũ tại: {output_be_path}")
        workbook.save(output_be_path)
        print(f"\nDEBUG: Đã xuất file Excel thành công tới: {output_be_path}")
        return output_be_path
    except Exception as e:
        print(f"DEBUG ERROR: Lỗi khi lưu file Excel: {str(e)}")
        # Ném lại lỗi để hàm gọi có thể bắt được và xử lý
        raise e

def merge_slot_cells(sheet: openpyxl.worksheet.worksheet.Worksheet, start_row: int) -> None:
    """
    Merges cells in the 'Time Slot' column that have the same value.

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The Excel sheet to process.
        start_row (int): The starting row of the data table.
    """
    if sheet.max_row < start_row:
        return

    current_slot = None
    merge_start = start_row
    
    for row in range(start_row, sheet.max_row + 2):
        slot_value = sheet.cell(row=row, column=1).value
        
        if slot_value != current_slot:
            if current_slot is not None and row - merge_start > 1:
                sheet.merge_cells(start_row=merge_start, start_column=1, end_row=row - 1, end_column=1)
            merge_start = row
            current_slot = slot_value